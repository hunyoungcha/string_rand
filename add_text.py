#hwp의 받아온 text를 필요한 부분만 정리해서 all_data.txt에 저장하는 코드

import os
import hwp_text as ht
from openpyxl import load_workbook
from openpyxl import Workbook

#파일 위치 및 리스트
path="rhksckf/"
file_list=os.listdir(path)


g=[]



#hwp의 받아온 텍스트를 name obs note의 프레임으로 만들어 temp.txt에 저장
def add_temp():
    stop=0
    temp=open('temp.txt', 'w', encoding='utf-8')
    for q in file_list:
        hwp_text=ht.get_hwp_text(path+q)
        f=hwp_text.split('\n')

        for i in range(5,len(f)):
            g=f[i].replace(' ','')
            if g.isspace()==False :
                if '공결' in g or '코로나' in g or '(' in g:
                    temp.write(g)
                    temp.write('공결\n')
                    i+=1 
                else:
                    temp.write(g)
    temp.close()

#temp 에서 엑셀로 옮기는 함수
def temp_to_xslx():
    #엑셀 열기
    wb=load_workbook('database.xlsx')
    ws=wb['database']
    #temp.txt 열기
    t=open('temp.txt','r', encoding='utf-8')
    data_list=t.readlines()
    t.close()

    cnt=0
    for i in range(1,(len(data_list)//3)+1): #row
        for j in range(1,4): #column   
            if data_list[i][0].isalpha():
                while True:
                    ws.cell(row=i,column=j,value=data_list[cnt].split('\n')[0])
                    cnt+=1
                    break
    wb.save('database.xlsx')

add_temp()