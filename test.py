from openpyxl import load_workbook
from openpyxl import Workbook

wb=load_workbook('database.xlsx')
ws=wb['database']

ws['A1']='hello'
ws['A2']='world'

wb.save("database.xlsx")

